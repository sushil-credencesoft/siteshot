# runner logic placeholder (use earlier provided runner code here)
#!/usr/bin/env python3
"""
Universal Test Runner for site_shot.py
======================================

Supports:
- CSV
- XLSX
- JSON
- XML

Generates:
- execution_report.csv
- per-test stdout / stderr logs

USAGE
-----
python run_from_file.py --file site_shot_test_cases.csv
python run_from_file.py --file site_shot_test_cases.xlsx
python run_from_file.py --file site_shot_test_cases.json
python run_from_file.py --file site_shot_test_cases.xml
"""

import argparse
import csv
import json
import os
import subprocess
import time
from datetime import datetime
from xml.etree import ElementTree as ET

try:
    import openpyxl
except ImportError:
    openpyxl = None


# -------------------------------------------------
# Helpers
# -------------------------------------------------

def utc_now():
    return datetime.utcnow().isoformat()


def safe_name(text: str):
    return "".join(c if c.isalnum() else "_" for c in text)[:60]


def run_command(command: str):
    start = time.time()
    proc = subprocess.Popen(
        command,
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True
    )
    stdout, stderr = proc.communicate()
    end = time.time()

    return {
        "exit_code": proc.returncode,
        "stdout": stdout,
        "stderr": stderr,
        "duration": round(end - start, 2)
    }


# -------------------------------------------------
# Loaders
# -------------------------------------------------

def load_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_xlsx(path):
    if not openpyxl:
        raise RuntimeError("openpyxl required for XLSX support")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cases.append(dict(zip(headers, row)))
    return cases


def load_xml(path):
    tree = ET.parse(path)
    root = tree.getroot()
    cases = []
    for tc in root.findall("testcase"):
        cases.append({child.tag: child.text for child in tc})
    return cases


def load_test_cases(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return load_csv(path)
    if ext == ".json":
        return load_json(path)
    if ext == ".xlsx":
        return load_xlsx(path)
    if ext == ".xml":
        return load_xml(path)
    raise ValueError(f"Unsupported file type: {ext}")


# -------------------------------------------------
# Main
# -------------------------------------------------

def main():
    parser = argparse.ArgumentParser("Universal QA Test Runner")
    parser.add_argument("--file", required=True, help="CSV/XLSX/JSON/XML test file")
    parser.add_argument("--out-dir", default="test_results")
    args = parser.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)
    logs_dir = os.path.join(args.out_dir, "logs")
    os.makedirs(logs_dir, exist_ok=True)

    report_path = os.path.join(args.out_dir, "execution_report.csv")

    test_cases = load_test_cases(args.file)

    report_fields = [
        "Test_ID",
        "Test_Title",
        "Command",
        "Start_Time",
        "End_Time",
        "Duration_sec",
        "Exit_Code",
        "Status",
        "Stdout_Log",
        "Stderr_Log"
    ]

    with open(report_path, "w", newline="", encoding="utf-8") as report:
        writer = csv.DictWriter(report, fieldnames=report_fields)
        writer.writeheader()

        for idx, tc in enumerate(test_cases, start=1):
            test_id = tc.get("Test_ID", f"TC-{idx}")
            title = tc.get("Test_Title", "N/A")
            command = tc.get("Command")

            if not command:
                continue

            print(f"\nâ–¶ {test_id}: {title}")
            print(f"  CMD: {command}")

            start = utc_now()
            result = run_command(command)
            end = utc_now()

            base = safe_name(f"{test_id}_{title}")
            stdout_file = os.path.join(logs_dir, f"{base}_stdout.log")
            stderr_file = os.path.join(logs_dir, f"{base}_stderr.log")

            with open(stdout_file, "w", encoding="utf-8") as f:
                f.write(result["stdout"])
            with open(stderr_file, "w", encoding="utf-8") as f:
                f.write(result["stderr"])

            status = "PASS" if result["exit_code"] == 0 else "FAIL"

            writer.writerow({
                "Test_ID": test_id,
                "Test_Title": title,
                "Command": command,
                "Start_Time": start,
                "End_Time": end,
                "Duration_sec": result["duration"],
                "Exit_Code": result["exit_code"],
                "Status": status,
                "Stdout_Log": stdout_file,
                "Stderr_Log": stderr_file
            })

            print(f"  âœ” {status} ({result['duration']}s)")

    print("\nâœ… All test cases executed")
    print(f"ðŸ“„ Report generated: {report_path}")


if __name__ == "__main__":
    main()
